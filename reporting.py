import os
import json
import shutil
import openpyxl
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.worksheet.properties import PageSetupProperties
from datetime import datetime

class ReportMaker:

    def __init__(self, template_path):
        self.template_path = template_path

    @staticmethod
    def copy_file(src_path, dst_path):
        try:
            shutil.copy2(src_path, dst_path)
        except IOError as e:
            print(f"파일 복사 오류: {e}")

    def create_report(self, output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
            
        with open('conditions.json', 'r') as file:
            report = json.load(file)
        with open('calculation_raw.json', 'r') as file:
            report.update(json.load(file)["report"])

        # make float values into string
        for i, j in report.items():
            if isinstance(j, float):
                if i == "r^2-" or i == "r^2+":
                    report[i] = f"{j:.4f}"
                else:
                    report[i] = f"{j:.2f}"

        now = datetime.now().strftime("%d%m%Y-%H%M%S")
        os.makedirs("reports", exist_ok=True)
        with open(f'./reports/report_{now}.json', 'w') as file:
            json.dump(report, file, indent=4)

        for row in ws.iter_rows():
            for cell in row:
                try:
                    report_item = cell.value.split("*")
                    if item := report.get(report_item[1]):
                        cell.value = item
                    else:
                        cell.value = "-"
                except IndexError:
                    pass
                except AttributeError:
                    pass

        wb.save(output_path)
        wb.close()

    @staticmethod
    def insert_image(file_path, image_path, cell, width=None, height=None):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        img = XlsxImage(image_path)
        img.width = width if width else img.width
        img.height = height if height else img.height
        ws.add_image(img, cell)

        wb.save(file_path)
        wb.close()

    @staticmethod
    def fit_to_a4(file_path, image_cell="B28", image_height=355):
        """PDF 변환 시 A4 세로 한 장에 모두 들어가도록 페이지를 설정한다.

        템플릿에는 용지·배율 설정이 없어 기본값으로 변환되면 표와 그래프가
        여러 장으로 잘린다. 여백을 줄이고 fitToPage 로 한 장에 맞춘다.
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        # fitToPage 는 sheet_properties 쪽 플래그가 켜져 있어야 실제로 적용된다
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        # 여백을 줄여 인쇄 영역을 최대한 확보한다 (단위: 인치)
        ws.page_margins.left = ws.page_margins.right = 0.25
        ws.page_margins.top = ws.page_margins.bottom = 0.3
        ws.page_margins.header = ws.page_margins.footer = 0.1

        # 인쇄 영역: 표(A1:E27) + 아래에 삽입된 그래프까지 포함해야 한다.
        # 이미지 높이(px)를 행 높이(기본 20px)로 환산해 마지막 행을 구한다.
        first_row = int(''.join(c for c in image_cell if c.isdigit()))
        last_row = first_row + (image_height // 20) + 2
        last_row = max(last_row, ws.max_row)
        ws.print_area = f"A1:E{last_row}"
        ws.page_setup.horizontalCentered = True

        wb.save(file_path)
        wb.close()

    @staticmethod
    def protect_excel_file(file_path, key):
        wb = openpyxl.load_workbook(file_path)
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.protection.sheet = True
            ws.protection.password = key
        
        wb.save(file_path)
        wb.close()
    
    # @staticmethod    
    # def xlsx_to_image(input_file, output_file):
    #     # Load the Excel workbook
    #     workbook = openpyxl.load_workbook(filename=input_file)
    #     # Select the first sheet or specify the sheet name if needed
    #     sheet = workbook.active
    #     # Convert the sheet to an image using pyvips
    #     image = pyvips.Image.new_from_array(sheet.iter_rows(values_only=True))
    #     # Save the image to a file
    #     image.write_to_file(output_file)

    def process_report(self, output_path, image_path, cell, width, height, key):

        self.copy_file(self.template_path, output_path)
        self.create_report(output_path)
        self.insert_image(output_path, image_path, cell, width, height)
        # 시트 보호 전에 페이지 설정을 끝내야 한다
        self.fit_to_a4(output_path, cell, height)
        self.protect_excel_file(output_path, key)
        # self.xlsx_to_image(output_path, output_path.split(".")[0] + ".jpg")

if __name__ == "__main__":
    # 템플릿 파일 이름
    template_path = "report_template.xlsx"
    # 결과 파일 명
    now = datetime.now().strftime("%y%m%d-%H%M%S")
    output_path = f"report_{now}.xlsx"
    # 이미지 정보
    image_path = "graph.png"
    cell = "B28"
    width = 590
    height = 355
    # 파일 보호 비밀번호
    key = "asdf"
    report_maker = ReportMaker(template_path)
    report_maker.process_report(output_path, image_path, cell, width, height, key)

    import platform
    if platform.system() == "Linux":
        import subprocess as sub
        sub.call(f"libreoffice --headless --convert-to pdf {output_path}", shell=True)
        shutil.move(output_path.split('.')[0] + '.pdf', "report.pdf")
        sub.call(f"xpdf ./report.pdf", shell=True)
